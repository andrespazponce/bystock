"""
Importador de estados financieros desde Excel (.xlsx / .xls).

Formato soportado (dos filas de encabezado):

    Col A          Col B         Col C         Col D         Col E
    (vacío)        Enero 2021    Enero 2021    Febrero 2021  Febrero 2021
    (vacío)        Incerpaz      EmpresaB      Incerpaz      EmpresaB
    Activo Total   1,500,000     800,000       1,600,000     850,000
    Activo Cte       900,000     400,000         950,000     420,000
    ...

Reglas de detección:
  · Fila de períodos: primera fila (dentro de las 20 primeras) que tiene
    al menos un valor parseable como período en columna >= 1.
    Los períodos se "rellenan hacia la derecha" para manejar celdas fusionadas
    (merged cells exportadas como null en las celdas secundarias).
  · Fila de empresas: primera fila DESPUÉS de la fila de períodos que tiene
    al menos un nombre de empresa reconocido en columna >= 1.
  · Compatibilidad hacia atrás: si el período está en cualquier celda
    (incluso col A o una fila suelta) y no hay "fila de períodos" en col >= 1,
    se usa ese período como global para todas las columnas.

Cada columna de datos (B, C, D…) queda mapeada a un par (Empresa, Período).
Se crea un PeriodoImport por par único y un ValorCuenta por celda.
Reimportación: elimina los ValorCuenta anteriores del período antes de insertar.
"""

import re
from decimal import Decimal, InvalidOperation

import openpyxl

from core.models import Empresa

from .models import CuentaContable, PeriodoImport, ValorCuenta

# ── Meses en inglés y español → número ───────────────────────────────────────

_MESES = {
    'january': 1,  'febrero': 2,
    'february': 2, 'febrero': 2,
    'march': 3,    'marzo': 3,
    'april': 4,    'abril': 4,
    'may': 5,      'mayo': 5,
    'june': 6,     'junio': 6,
    'july': 7,     'julio': 7,
    'august': 8,   'agosto': 8,
    'september': 9,'septiembre': 9,
    'october': 10, 'octubre': 10,
    'november': 11,'noviembre': 11,
    'december': 12,'diciembre': 12,
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
    'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9,
    'oct': 10,'nov': 11,'dec': 12,
    'ene': 1, 'abr': 4, 'ago': 8,
}


def _parsear_periodo(valor):
    """
    Extrae (mes, anio) de un valor de celda.
    Acepta: datetime/date, "January 2021", "Enero 2021", "01/2021", "2021-01-31".
    Devuelve (mes:int, anio:int) o (None, None).
    """
    if valor is None:
        return None, None

    # Objeto fecha nativo de openpyxl/Excel
    if hasattr(valor, 'month') and hasattr(valor, 'year'):
        return valor.month, valor.year

    texto = str(valor).strip()
    partes = re.split(r'[\s,/\-]+', texto.lower())

    if len(partes) == 2:
        a, b = partes
        if a in _MESES and b.isdigit() and len(b) == 4:
            return _MESES[a], int(b)
        if b in _MESES and a.isdigit() and len(a) == 4:
            return _MESES[b], int(a)

    m = re.match(r'^(\d{1,2})[/\-](\d{4})$', texto)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.match(r'^(\d{4})[/\-](\d{1,2})$', texto)
    if m:
        return int(m.group(2)), int(m.group(1))

    return None, None


def _parsear_valor(celda_valor):
    """Convierte el valor de una celda numérica a Decimal o None."""
    if celda_valor is None or celda_valor == '':
        return None
    if isinstance(celda_valor, (int, float)):
        return Decimal(str(celda_valor))
    texto = str(celda_valor).strip().replace(',', '').replace(' ', '')
    if texto.startswith('(') and texto.endswith(')'):
        texto = '-' + texto[1:-1]
    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def importar_balance(archivo, tipo_estado='BG'):
    """
    Lee el Excel y guarda los datos estructurados en la base de datos.

    Devuelve un dict con:
        importados      → total de ValorCuenta creados
        periodos        → lista de labels de períodos procesados ("Enero 2021", …)
        empresas        → lista de {nombre, importados} por empresa
        no_encontrados  → nombres de cuenta del Excel no reconocidos
        error           → string si hubo error fatal (None si todo OK)
    """
    resultado = {
        'importados':     0,
        'periodos':       [],
        'empresas':       [],
        'no_encontrados': [],
        'error':          None,
    }

    # ── 1. Abrir el workbook ─────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
    except Exception as exc:
        resultado['error'] = f'No se pudo abrir el archivo: {exc}'
        return resultado

    ws = wb.active

    # ── 2. Pre-cargar empresas activas ───────────────────────────────────────
    empresas_map = {}
    for e in Empresa.objects.filter(activa=True):
        empresas_map[e.nombre.lower().strip()] = e
        empresas_map[e.codigo.lower().strip()] = e

    # ── 3. Leer las primeras 20 filas para detectar encabezados ─────────────
    header_rows = list(ws.iter_rows(max_row=20, values_only=True))
    max_col = max((len(r) for r in header_rows), default=2)

    # ── 4. Detectar la fila de períodos ──────────────────────────────────────
    # Primera fila que tiene al menos un período parseado en columna >= 1.
    period_row_idx = None
    for row_idx, row in enumerate(header_rows):
        for col_idx in range(1, len(row)):
            m, a = _parsear_periodo(row[col_idx])
            if m and a:
                period_row_idx = row_idx
                break
        if period_row_idx is not None:
            break

    # ── 5. Rellenar períodos hacia la derecha (fill-forward) ─────────────────
    # Necesario para manejar celdas fusionadas: "Enero 2021" fusionado sobre
    # 5 columnas → openpyxl lee el valor solo en la primera, null en las demás.
    filled_periods = {}  # col_idx → (mes, anio)
    if period_row_idx is not None:
        period_row = header_rows[period_row_idx]
        current = None
        for col_idx in range(1, max_col):
            val = period_row[col_idx] if col_idx < len(period_row) else None
            if val is not None:
                m, a = _parsear_periodo(val)
                if m and a:
                    current = (m, a)
            if current:
                filled_periods[col_idx] = current

    # Fallback: buscar un período global en cualquier celda (formato legado)
    if not filled_periods:
        for row in header_rows:
            for val in row:
                m, a = _parsear_periodo(val)
                if m and a:
                    filled_periods = {col_idx: (m, a) for col_idx in range(1, max_col + 1)}
                    break
            if filled_periods:
                break

    # ── 6. Detectar la fila de empresas ──────────────────────────────────────
    # Primera fila DESPUÉS de la fila de períodos con empresas en col >= 1.
    empresa_row_idx = None
    search_from = (period_row_idx + 1) if period_row_idx is not None else 0
    for row_idx in range(search_from, len(header_rows)):
        row = header_rows[row_idx]
        for col_idx in range(1, len(row)):
            val = row[col_idx]
            if val is not None and str(val).strip().lower() in empresas_map:
                empresa_row_idx = row_idx
                break
        if empresa_row_idx is not None:
            break

    # Si no aparece después del período, buscar desde el principio
    if empresa_row_idx is None:
        for row_idx, row in enumerate(header_rows):
            for col_idx in range(1, len(row)):
                val = row[col_idx]
                if val is not None and str(val).strip().lower() in empresas_map:
                    empresa_row_idx = row_idx
                    break
            if empresa_row_idx is not None:
                break

    # ── 7. Construir col_mapping: col_idx → (Empresa, mes, anio) ─────────────
    col_mapping = {}
    if empresa_row_idx is not None:
        empresa_row = header_rows[empresa_row_idx]
        for col_idx in range(1, len(empresa_row)):
            val = empresa_row[col_idx]
            if val is None:
                continue
            key = str(val).strip().lower()
            if key in empresas_map:
                period = filled_periods.get(col_idx)
                if period:
                    col_mapping[col_idx] = (empresas_map[key], period[0], period[1])

    # ── 8. Validaciones ───────────────────────────────────────────────────────
    if not filled_periods:
        resultado['error'] = (
            'No se encontró ningún período en el archivo. '
            'Verificá que la fila con las fechas (ej. "Enero 2021") esté '
            'en las primeras 20 filas.'
        )
        return resultado

    if not col_mapping:
        resultado['error'] = (
            'No se encontraron empresas reconocidas en el encabezado. '
            'Verificá que los nombres coincidan exactamente con las empresas '
            'registradas en el sistema.'
        )
        return resultado

    # ── 9. Crear / reemplazar PeriodoImport por par (empresa, mes, anio) ─────
    periodos_cache = {}  # (empresa_id, mes, anio) → (PeriodoImport, Empresa)
    col_periodo    = {}  # col_idx → PeriodoImport

    for col_idx, (empresa, mes, anio) in col_mapping.items():
        cache_key = (empresa.id, mes, anio)
        if cache_key not in periodos_cache:
            periodo, creado = PeriodoImport.objects.get_or_create(
                empresa=empresa,
                anio=anio,
                mes=mes,
                tipo_estado=tipo_estado,
                defaults={'publicado': True},
            )
            if not creado:
                periodo.valores.all().delete()
            periodos_cache[cache_key] = (periodo, empresa)
        col_periodo[col_idx] = periodos_cache[cache_key][0]

    # ── 10. Mapa de cuentas contables ─────────────────────────────────────────
    cuentas_map = {
        c.nombre_excel.lower().strip(): c
        for c in CuentaContable.objects.filter(tipo_estado=tipo_estado, activa=True)
    }

    # ── 11. Leer filas de datos ───────────────────────────────────────────────
    # Las filas de datos empiezan en la fila siguiente al último encabezado.
    last_header = max(
        period_row_idx  if period_row_idx  is not None else -1,
        empresa_row_idx if empresa_row_idx is not None else -1,
    )
    data_min_row = last_header + 2   # +1 para 1-based, +1 para saltar el header

    valores_a_crear = []
    ya_insertados   = set()   # (periodo_id, cuenta_id) — evita duplicados
    no_enc_vistos   = set()

    for row in ws.iter_rows(min_row=data_min_row, values_only=True):
        if not row or row[0] is None:
            continue
        nombre_key = str(row[0]).strip().lower()
        if not nombre_key:
            continue

        if nombre_key in cuentas_map:
            cuenta = cuentas_map[nombre_key]
            for col_idx, periodo in col_periodo.items():
                insert_key = (periodo.id, cuenta.id)
                if insert_key in ya_insertados:
                    continue
                ya_insertados.add(insert_key)
                val_celda = row[col_idx] if col_idx < len(row) else None
                valores_a_crear.append(
                    ValorCuenta(periodo=periodo, cuenta=cuenta, valor=_parsear_valor(val_celda))
                )
        elif len(nombre_key) > 3 and nombre_key not in no_enc_vistos:
            no_enc_vistos.add(nombre_key)
            resultado['no_encontrados'].append(str(row[0]).strip())

    # ── 12. Guardar en la BD ──────────────────────────────────────────────────
    ValorCuenta.objects.bulk_create(valores_a_crear)
    resultado['importados'] = len(valores_a_crear)

    # ── 13. Construir resumen por empresa ─────────────────────────────────────
    # Mapa periodo_id → Empresa
    periodo_to_empresa = {
        periodo.id: empresa
        for (periodo, empresa) in periodos_cache.values()
    }
    empresa_counts = {}   # empresa.id → {nombre, importados}
    for v in valores_a_crear:
        empresa = periodo_to_empresa.get(v.periodo_id)
        if empresa:
            if empresa.id not in empresa_counts:
                empresa_counts[empresa.id] = {'nombre': empresa.nombre, 'importados': 0}
            empresa_counts[empresa.id]['importados'] += 1

    resultado['empresas'] = list(empresa_counts.values())

    # ── 14. Lista de períodos procesados (ordenados cronológicamente) ─────────
    periodos_unicos = sorted(
        {(mes, anio) for (_, mes, anio) in periodos_cache.keys()},
        key=lambda x: (x[1], x[0])   # ordenar por año, luego mes
    )
    resultado['periodos'] = [
        PeriodoImport.MESES_ES[mes] + f' {anio}'
        for mes, anio in periodos_unicos
    ]

    # ── 15. Guardar archivo como referencia en el primer período ──────────────
    if hasattr(archivo, 'name') and periodos_cache:
        primer_periodo = next(iter(periodos_cache.values()))[0]
        primer_periodo.archivo = archivo
        primer_periodo.save(update_fields=['archivo'])

    return resultado
